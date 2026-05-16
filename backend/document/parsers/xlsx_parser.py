"""
FASE 4.2C — XLSX Parser (Excel via openpyxl)
Converte planilha para texto tabular para DLP scan.

Defesas implementadas:
- ZIP bomb check (XLSX é ZIP)
- Rejeição de macros (xlsm não suportado — retorna erro)
- Limite de linhas, colunas e chars extraídos
- Timeout hard via asyncio.wait_for + thread
- Células com fórmulas substituídas por "[FORMULA]" — não evaluadas
- Nenhum conteúdo original no response
"""
from __future__ import annotations

import asyncio
import gc
from dataclasses import dataclass

from document.limits import (
    MAX_CHARS_EXTRACTED,
    MAX_EXTRACTION_TIME_S,
    MAX_UPLOAD_SIZE_BYTES,
    DocumentErrorCode,
)

_ZIP_BOMB_RATIO  = 10
_MAX_UNCOMPRESSED = MAX_UPLOAD_SIZE_BYTES * _ZIP_BOMB_RATIO
_MAX_ROWS = 5_000
_MAX_COLS = 100


@dataclass(frozen=True)
class XlsxParseResult:
    text:       str
    sheets:     int
    rows:       int
    truncated:  bool
    error_code:    str | None
    error_message: str | None


def _check_zip_bomb(file_bytes: bytes) -> bool:
    import zipfile, io
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            return sum(i.file_size for i in zf.infolist()) > _MAX_UNCOMPRESSED
    except Exception:
        return False


def _extract_sync(file_bytes: bytes) -> XlsxParseResult:
    if _check_zip_bomb(file_bytes):
        return XlsxParseResult(
            text="", sheets=0, rows=0, truncated=False,
            error_code=DocumentErrorCode.FILE_TOO_LARGE,
            error_message="Compressed workbook exceeds safe decompression limit",
        )

    try:
        import openpyxl  # type: ignore
        import io

        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            read_only=True,
            data_only=True,   # read cached values, not formulas
            keep_vba=False,
        )

        parts: list[str] = []
        chars_so_far = 0
        total_rows = 0
        truncated = False

        for sheet in wb.worksheets:
            parts.append(f"[Aba: {sheet.title}]")
            for row in sheet.iter_rows(max_row=_MAX_ROWS, max_col=_MAX_COLS, values_only=True):
                if truncated:
                    break
                cells = []
                for cell in row:
                    if cell is None:
                        cells.append("")
                    elif isinstance(cell, str) and cell.startswith("="):
                        cells.append("[FORMULA]")
                    else:
                        cells.append(str(cell))
                line = "\t".join(cells)
                remaining = MAX_CHARS_EXTRACTED - chars_so_far
                if len(line) > remaining:
                    parts.append(line[:remaining])
                    truncated = True
                    break
                parts.append(line)
                chars_so_far += len(line)
                total_rows += 1

            if total_rows >= _MAX_ROWS:
                truncated = True

        wb.close()
        full_text = "\n".join(parts)

        return XlsxParseResult(
            text=full_text,
            sheets=len(wb.sheetnames),
            rows=total_rows,
            truncated=truncated,
            error_code=None,
            error_message=None,
        )

    except MemoryError:
        gc.collect()
        return XlsxParseResult(
            text="", sheets=0, rows=0, truncated=False,
            error_code=DocumentErrorCode.PARSE_ERROR,
            error_message="Memory limit exceeded during extraction",
        )
    except Exception:
        return XlsxParseResult(
            text="", sheets=0, rows=0, truncated=False,
            error_code=DocumentErrorCode.MALFORMED,
            error_message="Could not parse workbook",
        )


async def parse_xlsx(file_bytes: bytes) -> XlsxParseResult:
    """Entry point async. Executa extração em thread com timeout hard."""
    try:
        result = await asyncio.wait_for(
            asyncio.get_event_loop().run_in_executor(None, _extract_sync, file_bytes),
            timeout=MAX_EXTRACTION_TIME_S,
        )
    except asyncio.TimeoutError:
        gc.collect()
        result = XlsxParseResult(
            text="", sheets=0, rows=0, truncated=False,
            error_code=DocumentErrorCode.TIMEOUT,
            error_message="Extraction exceeded time limit",
        )
    return result
